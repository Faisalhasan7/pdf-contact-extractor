"""
PDF Contact Extractor
---------------------
Scans all PDF files in a folder and extracts Name, Email, and Phone Number.
Results are saved to contacts_output.xlsx and contacts_output.csv

Usage:
    pip install pdfplumber openpyxl
    python pdf_contact_extractor.py

By default, it looks for PDFs in the same folder as this script.
Change PDF_FOLDER below if your PDFs are somewhere else.
"""

import re
import os
import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path

# ── CONFIG ──────────────────────────────────────────────────────────────────
PDF_FOLDER = "."          # Folder containing your PDFs. Change this if needed.
OUTPUT_XLSX = "contacts_output.xlsx"
OUTPUT_CSV  = "contacts_output.csv"
# ────────────────────────────────────────────────────────────────────────────


# ── REGEX PATTERNS ───────────────────────────────────────────────────────────
EMAIL_PATTERN = re.compile(
    r'[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}'
)

PHONE_PATTERN = re.compile(
    r'(?:'
    r'\+?(\d{1,3})[\s\-.]?'   # optional country code
    r')?'
    r'[\(\[]?(\d{2,4})[\)\]]?' # area code
    r'[\s\-.]?'
    r'(\d{3,4})'               # middle segment
    r'[\s\-.]?'
    r'(\d{3,5})'               # last segment
)

# Name heuristic: looks for lines like "Name: John Doe" or standalone "First Last"
NAME_LABEL_PATTERN = re.compile(
    r'(?:name|full\s*name|candidate|applicant)[:\-\s]+([A-Z][a-zA-Z]+(?:\s[A-Z][a-zA-Z]+){1,3})',
    re.IGNORECASE
)
STANDALONE_NAME_PATTERN = re.compile(
    r'^([A-Z][a-z]{1,20}(?:\s[A-Z][a-z]{1,20}){1,2})\s*$',
    re.MULTILINE
)
# ─────────────────────────────────────────────────────────────────────────────


def extract_text_from_pdf(pdf_path):
    text = ""
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"
    except Exception as e:
        print(f"  [!] Could not read {pdf_path.name}: {e}")
    return text


def extract_email(text):
    matches = EMAIL_PATTERN.findall(text)
    return matches[0] if matches else ""


def extract_phone(text):
    matches = PHONE_PATTERN.findall(text)
    if not matches:
        return ""
    # Reconstruct the first valid match
    for m in matches:
        parts = [p for p in m if p]
        number = "".join(parts)
        if len(number) >= 7:
            return "".join(parts)
    return ""


def extract_name(text):
    # Try labelled name first (most reliable)
    m = NAME_LABEL_PATTERN.search(text)
    if m:
        return m.group(1).strip()

    # Fall back to standalone "First Last" on its own line
    # Usually near the top of a resume/form
    lines = text.strip().split("\n")
    for line in lines[:15]:  # Only check top 15 lines
        line = line.strip()
        m = STANDALONE_NAME_PATTERN.match(line)
        if m and len(line.split()) >= 2:
            return line
    return ""


def process_pdfs(folder):
    folder = Path(folder)
    pdf_files = sorted(folder.glob("*.pdf"))

    if not pdf_files:
        print(f"\n[!] No PDF files found in: {folder.resolve()}")
        print("    Make sure PDF_FOLDER points to the right directory.\n")
        return []

    print(f"\nFound {len(pdf_files)} PDF(s). Processing...\n")
    results = []

    for i, pdf_path in enumerate(pdf_files, 1):
        print(f"  [{i}/{len(pdf_files)}] {pdf_path.name}")
        text = extract_text_from_pdf(pdf_path)

        name  = extract_name(text)
        email = extract_email(text)
        phone = extract_phone(text)

        results.append({
            "File":  pdf_path.name,
            "Name":  name,
            "Email": email,
            "Phone": phone,
        })

        status = []
        if not name:  status.append("name missing")
        if not email: status.append("email missing")
        if not phone: status.append("phone missing")
        if status:
            print(f"         ⚠ {', '.join(status)}")

    return results


def save_xlsx(results, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Contacts"

    headers = ["File", "Name", "Email", "Phone"]
    header_fill = PatternFill("solid", start_color="4472C4")
    header_font = Font(bold=True, color="FFFFFF", name="Arial")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, record in enumerate(results, 2):
        ws.cell(row=row_idx, column=1, value=record["File"])
        ws.cell(row=row_idx, column=2, value=record["Name"])
        ws.cell(row=row_idx, column=3, value=record["Email"])
        ws.cell(row=row_idx, column=4, value=record["Phone"])
        # Subtle alternating rows
        if row_idx % 2 == 0:
            for col in range(1, 5):
                ws.cell(row=row_idx, column=col).fill = PatternFill("solid", start_color="DCE6F1")

    # Auto column widths
    col_widths = [40, 25, 35, 20]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF", name="Arial")

    wb.save(output_path)


def save_csv(results, output_path):
    import csv
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["File", "Name", "Email", "Phone"])
        writer.writeheader()
        writer.writerows(results)


def main():
    results = process_pdfs(PDF_FOLDER)
    if not results:
        return

    save_xlsx(results, OUTPUT_XLSX)
    save_csv(results, OUTPUT_CSV)

    found     = sum(1 for r in results if r["Name"] or r["Email"] or r["Phone"])
    complete  = sum(1 for r in results if r["Name"] and r["Email"] and r["Phone"])

    print(f"""
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
  Done! Processed {len(results)} PDF(s)
  ✔ Full data (name+email+phone): {complete}
  ⚠ Partial data:                 {found - complete}
  ✗ Nothing found:                {len(results) - found}

  Output saved to:
    → {OUTPUT_XLSX}
    → {OUTPUT_CSV}
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
""")


if __name__ == "__main__":
    main()
